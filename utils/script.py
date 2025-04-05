from itertools import chain
import json
import tqdm
import chromadb
import re
import pandas as pd
import numpy as np
import ast
import sys
import os
os.environ["TOKENIZERS_PARALLELISM"] = "false"


import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from utils.cal_embedding_bge_zh import calculate_docs_embedding_zh, get_embeddings_zh
from utils.load_source_excel import get_project_df, get_industry_coop_proj
from langchain_community.vectorstores.chroma import Chroma
from utils.load_former_manager import get_former_manager

from utils.get_setting import setting_data, print_setting_data, find_key_path, value_of_key
from utils.filter_method import *

class MissingFieldsException(Exception):
    pass

def load_into_chroma_bge_manager(is_industry=False):
    # Blacklist of fields to be removed
    blacklist_csv_path = find_key_path("退休或黑名單委員")
    blacklist_df = pd.read_csv(blacklist_csv_path, encoding='utf-8')
    blacklist = blacklist_df['姓名'].tolist()
    print(f"黑名單或是已退休委員:", blacklist)
    
    # Determine the Chroma database path based on the industry flag
    chroma_db_key = 'CHROMA_INDUSTRY' if is_industry else 'CHROMA'
    chroma_db_path = find_key_path(chroma_db_key)
    client = chromadb.PersistentClient(path=chroma_db_path)
    collection_name = chroma_db_key
    
    client.delete_collection(collection_name)
    collection = client.create_collection(collection_name)

    # Retrieve the relevant dataframes based on the industry flag
    df_list = get_industry_coop_proj() if is_industry else get_project_df()

    manager_group = {}
    for key, year_data in df_list.items():
        for i in tqdm.tqdm(range(len(year_data)), desc=key):
            row = year_data.iloc[i]
            
            manager = year_data.iloc[i]['計畫主持人']
            project_name = row.get('計畫中文名稱', '') or ''
            abstract = row.get('中文摘要', '') or ''
            keywords = row.get('中文關鍵字', '') or ''
            approved = str(row.get('通過', 'false')).strip().lower()
            
            # Skip projects that are not approved when not dealing with industry data
            if not is_industry and approved != 'true':  
                continue
            
            # Skip managers in the blacklist
            if manager in blacklist:
                continue
            
            # Concatenate the project information
            text = f"{project_name} {abstract} {keywords}\n"
            
            # Group by manager, appending project information
            if manager in manager_group:
                manager_group[manager] += f"\n{text}"
            else:
                manager_group[manager] = text

    # Save the manager data into the Chroma database
    for manager, text in tqdm.tqdm(manager_group.items(), desc="Saving to Chroma"):
        # Attempt to calculate embeddings with a max of 3 retries
        embeddings = None
        for _ in range(3):
            embeddings = calculate_docs_embedding_zh([text])
            if embeddings:
                break
        
        # Upsert manager's project data into the Chroma collection
        collection.upsert(
            documents=[text],
            ids=[manager],
            embeddings=embeddings,
            metadatas=[{'manager': manager}]
        )

    # Determine the path for saving manager group data
    bge_manager_key = "BGE_INDUSTRY_MANAGER" if is_industry else "BGE_MANAGER"
    bge_manager_path = find_key_path(bge_manager_key)
    
    # Save manager group data to a JSON file
    with open(bge_manager_path, 'w', encoding='utf-8') as f:
        json.dump(manager_group, f, ensure_ascii=False)



def search_v3(is_industry=False):
    
    # 取得計畫相關的欄位值
    tabs = value_of_key("計畫SHEET")

    # 取得輸出 Excel 檔案的資料夾路徑
    output_excel_folder_path = find_key_path("統計表分析")

    # 判斷「研究計畫」與「產學合作」的資料庫路徑
    if is_industry:
        chroma_db_path = find_key_path('CHROMA_INDUSTRY')
        vectorstore = Chroma("CHROMA_INDUSTRY", persist_directory=chroma_db_path, embedding_function=get_embeddings_zh())
        excel_folder_path = find_key_path("產學合作申請名冊")
    else:
        chroma_db_path = find_key_path('CHROMA')
        vectorstore = Chroma("CHROMA", persist_directory=chroma_db_path, embedding_function=get_embeddings_zh())
        excel_folder_path = find_key_path("研究計畫申請名冊")
        
    former_manager = get_former_manager(find_key_path("曾任委員"))
    
    # 要輸出的欄位
    other_fields = value_of_key("計畫相關其他欄位")
    required_fields = []
    keys = [
        "計畫名稱",
        "中文關鍵字",
        "計劃摘要",
        "職稱",
        "申請主持人欄位名稱",
        "申請機構欄位名稱",
        
    ]
    for key in keys:
        temp_value = value_of_key(key)
        if temp_value is not None and temp_value != '':
            required_fields.append(temp_value)
        
    multi_keys = [
        "申請共同主持人"
    ]
    for key in multi_keys:
        temp_value = value_of_key(key)
        for value in temp_value:
            if value not in required_fields and value != '':
                required_fields.append(value)
    
    filter_fields = required_fields + [f for f in other_fields if f not in required_fields]

    RECOMMAND_AMOUNT = 10   # 要推薦的委員數量
    SELECT_AMOUNT = 3       # 要選擇的委員數量
    SELECT_BOX_SYMBOL = ['Y', 'Z', 'AA']

    xls = pd.ExcelFile(excel_folder_path)
    writer = pd.ExcelWriter(output_excel_folder_path, engine='openpyxl')
    
    similarity_record_path = f"./data/output/similarity_record_{value_of_key('FINAL_COMMITTEE')}"
    os.makedirs(os.path.dirname(similarity_record_path), exist_ok=True)
    similarity_df = pd.DataFrame(columns=["query_text", "compared_text", "recommended_manager", "model_name", "similarity_score"])
    
    try:
        project_name_field_name = value_of_key("計畫名稱")
        chinese_keyword_field_name = value_of_key("中文關鍵字")
        abstract_field_name = value_of_key("計劃摘要")
        
        for tab in tabs:
            page_manager_list = []

            # define column name
            df = pd.read_excel(xls, tab)
            
            # 檢查 filter_fields 是否在現有的欄位中
            existing_fields = df.columns.tolist()
            missing_fields = [field for field in filter_fields if field not in existing_fields]
            if missing_fields:
                print("現有欄位:", existing_fields)
                print("應當欄位:", filter_fields)
                raise ValueError("欄位不匹配，程式碼運行停止")  # 引發例外，中止程式碼運行

            df = df[filter_fields]

            for i in range(RECOMMAND_AMOUNT):
                df['推薦委員' + str(i + 1)] = ''
                df['相關分數' + str(i + 1)] = ''
            df['前任委員占比'] = ''
            for i in range(SELECT_AMOUNT):
                df['選取委員' + str(i + 1)] = ''

            # process data
            for i in tqdm.tqdm(range(len(df)), desc=tab):
                manager_list = []
                project_name = df.iloc[i].get(project_name_field_name, '')
                keywords = df.iloc[i].get(chinese_keyword_field_name, '')
                abstract = df.iloc[i].get(abstract_field_name, '')
                
                # 找尋相似度
                current_text_combine = f"{project_name} {keywords} {abstract}"
                documents = vectorstore.similarity_search_with_relevance_scores(
                    current_text_combine,
                    k=RECOMMAND_AMOUNT
                )
                
                # 將搜尋結果寫入 CSV
                for doc, score in documents:
                    recommended_manager = doc.metadata['manager'] 
                    compared_text = doc.page_content  # 可根據需求選擇不同內容
                    model_name = "BGE_ZH"  # 依實際使用的模型名稱
                    
                    # 追加數據
                    new_row = pd.DataFrame([{
                        "query_text": current_text_combine,
                        "compared_text": compared_text,
                        "recommended_manager": recommended_manager,
                        "model_name": model_name,
                        "similarity_score": score
                    }])
                    
                    new_row = new_row.reindex(columns=similarity_df.columns)
                    if not new_row.empty and not new_row.isna().all(axis=None):
                        similarity_df = pd.concat([similarity_df, new_row], ignore_index=True)

        
                # 分數填入 Excel 的動作 (和原程式邏輯相同)
                for j, (doc, score) in enumerate(documents):
                    df.loc[df.index[i], '推薦委員' + str(j + 1)] = doc.metadata['manager']
                    manager_list.append(doc.metadata['manager'])
                    df.loc[df.index[i], '相關分數' + str(j + 1)] = score

                page_manager_list.append(manager_list)
                df.loc[df.index[i], '前任委員占比'] = len([x for x in manager_list if x in former_manager]) / RECOMMAND_AMOUNT

            df.to_excel(writer, sheet_name=tab, index=False)

            # setup dropdown list
            workbook = writer.book
            worksheet = workbook[tab]

            for j in range(SELECT_AMOUNT):
                for i, manager_list in enumerate(page_manager_list):
                    data_range = ','.join(manager_list)
                    dv = DataValidation(type="list", formula1=f'"{data_range}"', allow_blank=True)
                    dv.add(SELECT_BOX_SYMBOL[j] + str(i + 2))
                    worksheet.add_data_validation(dv)

            highligh_former_manager(writer, tab, former_manager, output_excel_folder_path)
            draw_color_for_similarity_score(writer, tab, output_excel_folder_path)

    except Exception as e:
        if not writer.book.sheetnames:
            print("ERROR")
            writer.book.create_sheet(title="Error")
        raise  # 重新引發異常以停止程式

    finally:
        # 確保 ExcelWriter 正常關閉
        writer.close()  
        with pd.ExcelWriter(similarity_record_path, engine='openpyxl') as similarity_writer:
            similarity_df.to_excel(similarity_writer, sheet_name="Similarity Records", index=False)
            
            
def draw_color_for_similarity_score(writer, tab, output_excel):
    
    from openpyxl.formatting.rule import ColorScaleRule
    
    SIMILARITY_SCORE_RANGE = '$E$2:$w$1000'
    workbook = writer.book
    worksheet = workbook[tab]
    rule = ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="F9F900")
    worksheet.conditional_formatting.add(SIMILARITY_SCORE_RANGE, rule)
    workbook.save(output_excel)

def highligh_former_manager(writer, tab, former_manager, output_excel):
    
    from openpyxl.formatting import Rule
    from openpyxl.styles.differential import DifferentialStyle
    
    RECOMMAND_MANAGER_SYMBOL = ['D','F','H','J','L','N','P','R','T','V']
    workbook = writer.book
    worksheet = workbook[tab]
    redFill = PatternFill(start_color='FFA042', end_color='FFA042', fill_type='solid')

    for s in RECOMMAND_MANAGER_SYMBOL:
        col1 = worksheet[s]
        for i, cell in enumerate(col1):
            cell_value = cell.value
            if cell_value in former_manager:
                rule = Rule(type="cellIs", operator="equal", formula=[f'"{cell_value}"'], dxf=DifferentialStyle(fill=redFill))
                # rule = Rule(type="expression", operator="equal", formula=[f'"{cell_value}"'], dxf=DifferentialStyle(fill=redFill))
                worksheet.conditional_formatting.add(f'{s}{str(i+1)}', rule)

    workbook.save(output_excel)

def statistic_committee():
    
    apply_project_file_year = value_of_key("計畫過去申請案件年分範圍")
    
    statistic_folder_path = find_key_path("統計清單") 
    statistic_excel_file = pd.ExcelFile(statistic_folder_path)
    
    industry_folder_path = find_key_path("產學過去申請名冊")
    industry_data = pd.read_excel(industry_folder_path)
    
    past_apply_project_path = find_key_path("計畫過去申請案件")
    past_apply_project_file = pd.ExcelFile(past_apply_project_path)
    
    newest_person_path = find_key_path("暫存最新人才資料庫")
    newest_person_df = pd.read_excel(newest_person_path)
    
    # committee_study_folder_path = find_key_path("碩博士論文_RDF")
    # committee_study_data = pd.read_excel(committee_study_folder_path)
    
    #@ 處理委員的所有相關學校名單: 名稱 - 年份 - 學校 - 職稱
    committee_person_RDF = []
    
    # #: 碩博士論文
    # for index, row in tqdm.tqdm(committee_study_data.iterrows(), desc="碩博士論文"):
    #     # 先檢查欄位是否存在
    #     name = row.get('學生姓名', '')
    #     year = row.get('畢業學年度', '') if '畢業學年度' in row else ''
    #     institution = row.get('畢業學校', '') if '畢業學校' in row else ''

    #     # 確保數據不為 NaN（轉為字串或預設值）
    #     year = str(year) if pd.notna(year) else ''
    #     institution = str(institution) if pd.notna(institution) else ''

    #     committee_person_RDF.append({
    #         '名稱': name,
    #         '年份': year,
    #         '機關名稱': institution,
    #         '職稱': ""
    #     })
    
    #: 暫存最新人才資料庫
    for index, row in newest_person_df.iterrows():
        committee_person_RDF.append({
            '名稱': row['名稱'],
            '年份': int(row['年份']),
            '機關名稱': row['機關名稱'],
            '職稱': row['職稱'],
            "來源": row['來源']
        })

    #: 研究計劃（申請案件）
    for year in apply_project_file_year:
        current_sheet = year
        past_apply_project_df = pd.read_excel(past_apply_project_file, current_sheet)
        for index, row in tqdm.tqdm(past_apply_project_df.iterrows(), desc=f"{current_sheet}"):
            committee_person_RDF.append({
                '名稱': row['計畫主持人'],
                '年份': year,
                '機關名稱': row['機關名稱'],
                '職稱': row['職稱'],
                "來源": f"研究計劃（申請案件）- {current_sheet}"
            })
        
        
    #: 研究計劃（統計案件）
    for year in apply_project_file_year:
        current_sheet = f"{year}總計畫清單"
        statistic_df = pd.read_excel(statistic_excel_file, current_sheet)
        for index, row in tqdm.tqdm(statistic_df.iterrows(), desc=f"{current_sheet}"):
            committee_person_RDF.append({
                '名稱': row['計畫主持人'],
                '年份': year,
                '機關名稱': row['機關名稱'],
                '職稱': row['職稱'],
                "來源": f"研究計劃（統計案件）- {current_sheet}"
            })
            
    #: 產學合作
    for index, row in industry_data.iterrows():
        committee_person_RDF.append({
            '名稱': row['計畫主持人'],
            '年份': row["計畫編號"][:3] if not pd.isna(row["計畫編號"]) else "",
            '機關名稱': row['單位名稱'],
            '職稱': row['職稱'],
            "來源": f"產學合作 - 序號:{row['序號']}"
        })
    
    committee_person_RDF_df = pd.DataFrame(committee_person_RDF)
    committee_person_RDF_df[['學校', '系所']] = committee_person_RDF_df['機關名稱'].apply(split_institution)
    committee_person_RDF_df = committee_person_RDF_df.sort_values(by=["名稱"])
    committee_person_RDF_df.to_excel(find_key_path("統計清單人才資料_RDF"), index=False)
    
    from utils.filter_method import extract_max_year
    
    committee_person_RDF_df['年份'] = committee_person_RDF_df['年份'].apply(extract_max_year)
    committee_person_RDF_df['年份'] = committee_person_RDF_df['年份'].fillna(0) 
    committee_person_RDF_df['年份'] = committee_person_RDF_df['年份'].astype(int)
    
    unique_person_RDF_df = committee_person_RDF_df.loc[committee_person_RDF_df.groupby('名稱')['年份'].idxmax()]
    unique_person_RDF_df.to_excel(find_key_path("統計清單人才資料_RDF_UNI"), index=False)

def filter_committee(is_industry=False):
    
    #: Load the data
    crawler_RDF_folder_path = find_key_path("碩博士論文_RDF")
    crawler_RDF_data = pd.read_excel(crawler_RDF_folder_path)
    
    statistical_analysis_folder_path = find_key_path("統計表分析") 
    statistical_analysis_file = pd.ExcelFile(statistical_analysis_folder_path)
    
    if is_industry: apply_list_folder_path = find_key_path("產學合作申請名冊")
    else: apply_list_folder_path = find_key_path("研究計畫申請名冊") 
        
    apply_list_file = pd.ExcelFile(apply_list_folder_path)
    
    committee_person_path = find_key_path("統計清單人才資料_RDF_UNI")
    committee_person_RDF_df = pd.read_excel(committee_person_path)
    
    #- Strategy
    writer = pd.ExcelWriter(find_key_path("過濾相近後統計表"), engine='openpyxl')
    
    #@ 審查委員不能與計劃申請學校有關
    for sheet in statistical_analysis_file.sheet_names:
        current_sheet_statistical_excel_data = pd.read_excel(statistical_analysis_file, sheet_name=sheet)
        result_dict = []
        all_apply_members = current_sheet_statistical_excel_data[value_of_key("申請主持人欄位名稱")].to_list() # 所有申請人
        
        for index, statistical_row in current_sheet_statistical_excel_data.iterrows():
        # ~ 每個統計表的 row.
        
            # = 審查委員的背景
            committee_person_dict = []
            for index_of_committee in range(1, 11):
            # ~ 推薦委員 10 人
                
                # 委員過去待過的學校
                been_list = []
                find_temp_df = committee_person_RDF_df[committee_person_RDF_df["名稱"] == statistical_row[f'推薦委員{index_of_committee}']]
                for index, row in find_temp_df.iterrows(): 
                    been_list.append(row["學校"])
                been_list = list(set(been_list)) 
                
                # 委員過去畢業的學校
                graduate_list = []
                relate_school = find_crawler_person_relative_school(f'推薦委員{index_of_committee}', crawler_RDF_data)
                graduate_list.extend(list(set(relate_school)))
                
                # 委員職稱
                committee_person_dict.append({
                    "委員名稱": statistical_row[f'推薦委員{index_of_committee}'],
                    "委員曾就職學校": been_list,
                    "委員過去畢業學校": graduate_list,
                    "職稱": find_temp_df['職稱']
                })
                
            # = 申請學校 + 主持人學校 + 共同主持人學校
            '''
            如果有重疊，就會被歸類到 Filtered Members 裏
            反之則留在 Remaining Members
            至於 Filter Reasons 則紀錄每個被篩掉委員的具體原因
            '''
            total_committee_person_dict_result = {
                'Filtered Members': [],
                'Remaining Members': [],
                'Filter Reasons': {}
            }
            
            for sheet in apply_list_file.sheet_names: 
                current_sheet_apply_excel_data = pd.read_excel(apply_list_file, sheet_name=sheet)
                find_temp_df = current_sheet_apply_excel_data[
                    current_sheet_apply_excel_data[value_of_key("計畫名稱")] == statistical_row[value_of_key("計畫名稱")]
                ]
                
                for index, row in find_temp_df.iterrows(): 
                    joint_person_list = row.get(value_of_key("申請共同主持人"), pd.Series()).tolist()
                    joint_department_list = row.get(value_of_key("申請共同機構欄位名稱"), pd.Series()).tolist() 
                    common_person_dict = extract_text_in_parentheses(joint_person_list)
                    common_department_dict = extract_text_in_parentheses(joint_department_list)
                    
                    common_joint_list = common_person_dict + common_department_dict
                    
                    # 找到關聯性
                    project_manager_school = list([find_crawler_person_relative_school(name, crawler_RDF_data) for name, department in common_joint_list])
                    apply_school = {
                        "申請人職稱": row.get(value_of_key("職稱"), ''),
                        "計畫申請學校": split_institution(row.get(value_of_key("申請機構欄位名稱"), ''))[0], 
                        "共同計畫主持的學校": [split_institution(department)[0] for name, department in common_joint_list],
                        "計畫主持人過去畢業的學校": find_crawler_person_relative_school(value_of_key("申請主持人欄位名稱"), crawler_RDF_data),
                        "共同主持人過去畢業的學校": list(chain.from_iterable(project_manager_school)), 
                    }
                    
                    #~ 審查委員不能與計劃申請學校(包含共同主持人)有關
                    # print("apply_school\n", apply_school)
                    # print("committee_person_dict\n", committee_person_dict)
                    filter_pairs = [("計畫申請學校", "委員曾就職學校"), ("共同計畫主持的學校", "委員曾就職學校")]
                    TITLE_RESTRICTIONS = {
                        "助理教授": ["教授", "研究員"], # 助理教授不能審教授或研究員
                        "助研究員": ["教授", "研究員"] # 助研究員不能審教授或研究員
                    }
                    current_committee_person_dict_result = filter_committee_advanced(
                        apply_school, 
                        committee_person_dict, 
                        filter_pairs, 
                        all_apply_members, 
                        TITLE_RESTRICTIONS,
                        whether_to_execute_the_option= {
                            "是否過濾申請人": True if is_industry else False,
                            "是否過濾相同學校": True,
                            "是否過濾職稱": True,
                        }
                    )
                    total_committee_person_dict_result = merge_committee_advanced(total_committee_person_dict_result, current_committee_person_dict_result)
                    
                    
                if len(find_temp_df) > 0: break #= 找不到東西，跳掉
        
            #- Input Selector
            # final_committee_person_list = [item for item in committee_person_dict["Remaining Members"][:3]]
            # for index, name in enumerate(final_committee_person_list):
            #     statistical_row[f"選取委員{index+1}"] = name
            
            #- Reason
            statistical_row["篩掉人員"] = total_committee_person_dict_result["Filtered Members"]
            statistical_row["篩選原因"] = total_committee_person_dict_result["Filter Reasons"]
            
            result_dict.append(statistical_row)
            
        pd.DataFrame(result_dict).to_excel(writer, sheet_name=sheet, index=False)
    writer.close()

def load_data(file_path):
    """
        讀取 Excel 檔案並回傳 workbook 和 worksheet.
    """
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    return workbook, worksheet

def generate_letters_excel(start_index, gap, count):
    def index_to_excel_column(index):
        column = ""
        while index > 0:
            index -= 1  # 將索引調整為 0 基礎
            column = chr(index % 26 + 65) + column
            index //= 26
        return column

    letters = []
    for i in range(count):
        letter_index = start_index + i * gap
        letters.append(index_to_excel_column(letter_index))
    return letters

def add_comments(target_ws, data_ws, columns_to_comment):
    """
        在目標工作表上添加註釋
    """
    
    # columns_to_comment = ['D', 'F', 'H', 'J', 'L', 'N', 'P', 'R', 'T', 'V']  # 需要添加註釋的欄位

    # 建立名稱與詳細資訊的對應字典
    # ['名稱', '年份', '機關名稱', '職稱', '來源', '學校', '系所']
    name_to_details = {
    data_ws.cell(row=i, column=1).value: \
        f"名稱: {data_ws.cell(row=i, column=1).value}\n"
        f"年份: {data_ws.cell(row=i, column=2).value}\n"
        f"機關: {data_ws.cell(row=i, column=3).value}\n"
        f"職稱: {data_ws.cell(row=i, column=4).value}\n"
        # f"來源: {data_ws.cell(row=i, column=5).value}"
        for i in range(2, data_ws.max_row + 1)
    }
    
    headers = [cell.value for cell in data_ws[1]]  # 取得第一列的標題名稱
    source_index = headers.index('來源') + 1  # Excel 的索引從 1 開始
    print("來源欄位的索引位置:", source_index)
    print("欄位名稱:", headers)

    # 在每個指定欄位添加註釋
    for col in columns_to_comment:
        for cell in target_ws[col]:
            if cell.value in name_to_details:
                comment_text = name_to_details[cell.value]
                comment = openpyxl.comments.Comment(comment_text, "Python Script")
                comment.width = 350  # 設置寬度
                comment.height = 100  # 設置高度
                cell.comment = comment
                
def excel_process_VBA():
    from openpyxl.utils.cell import column_index_from_string
    from openpyxl.utils import get_column_letter

    gap = 2
    range_count = 10
    
    #: load the excel data.
    talent_workbook, talent_sheet = load_data(find_key_path("統計清單人才資料_RDF"))
    committee_workbook, committee_sheet = load_data(find_key_path("過濾相近後統計表"))
    out_count = committee_sheet.max_column - ( range_count * gap ) - 6
    
    start_index = out_count + 1 
    letter_index = generate_letters_excel(start_index, gap, range_count) # start_index = Excel 26 進制的索引
    # timeline = [start_index + i * gap for i in range(range_count)]  
    
    add_comments(committee_sheet, talent_sheet, columns_to_comment=letter_index)
    pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid') #= 填色
    
    header_value = committee_sheet.cell(row=1, column=start_index).value
    print("[起頭] 第 {} 欄之標題為：{}".format(start_index, header_value))

    # 檢查 AB, 滿足條件改色
    for row in committee_sheet.iter_rows(min_row=2, max_col=committee_sheet.max_column):
        filter_list = ast.literal_eval(row[-2].value)
        
        #- 若有重複的的部分進行圖色（篩選委員）
        for col_letter in letter_index:
            col_index = column_index_from_string(col_letter) - 1
            if row[col_index].value in filter_list:
                row[col_index].fill = pink_fill

    # 保存
    committee_workbook.save(find_key_path("FINAL_COMMITTEE"))

from datetime import datetime
def update_peronsal_info_database(is_industry=False):
    '''
        更新暫存最新人才資料庫
    '''
    
    # 讀取最新 Excel 檔案
    if is_industry:
        apply_list_folder_path = find_key_path("產學合作申請名冊")
        file_name = value_of_key("產學合作申請名冊")
    else:
        apply_list_folder_path = find_key_path("研究計畫申請名冊") 
        file_name = value_of_key("研究計畫申請名冊")  # 修正錯誤的 `file_name(file_name)`

    apply_list_file = pd.ExcelFile(apply_list_folder_path)
    
    # 讀取暫存人才資料庫，若不存在則建立新的 DataFrame
    personal_info_database_path = find_key_path("暫存最新人才資料庫")

    if os.path.exists(personal_info_database_path):
        personal_info_database = pd.read_excel(personal_info_database_path)
    else:
        personal_info_database = pd.DataFrame(columns=['名稱', '年份', '機關名稱', '職稱', '來源'])

    new_data = []

    for sheet in apply_list_file.sheet_names: 
        current_sheet_apply_excel_data = pd.read_excel(apply_list_file, sheet_name=sheet)
        for _, row in current_sheet_apply_excel_data.iterrows():
            new_row = {
                '名稱': row.get(value_of_key('申請主持人欄位名稱')), 
                '年份': datetime.now().year - 1911,  # 直接填入現在的年份
                '機關名稱': row.get(value_of_key('申請機構欄位名稱'), ''),
                '職稱': row.get(value_of_key('職稱'), ''),
                '來源': file_name,
            }
            new_data.append(new_row)

    # 將新數據轉換為 DataFrame
    new_data_df = pd.DataFrame(new_data)

    # 合併數據，並以 "名稱" 為主鍵，保留最新年份的資料
    updated_database = pd.concat([personal_info_database, new_data_df]).sort_values(by=['名稱', '年份'], ascending=[True, False])
    updated_database = updated_database.drop_duplicates(subset=['名稱'], keep='first')  # 只保留最新的年份

    # 儲存回 Excel
    updated_database.to_excel(personal_info_database_path, index=False)
