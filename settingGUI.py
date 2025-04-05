import sys
import os
from ruamel.yaml import YAML
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication,
    QDialog,
    QLabel,
    QComboBox,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QMessageBox,
    QFileDialog,
    QListWidget,
    QListWidgetItem,
    QAbstractItemView,
    QScrollArea,
    QWidget,
    QGridLayout
)
from PyQt6.QtCore import Qt

# 定義設定檔路徑
DEFAULT_SETTING_YAML = "./setting.yaml"

# 初始化 YAML 物件（保留引號等格式）
yaml = YAML()
yaml.preserve_quotes = True


class SelectProjectAimDialog(QDialog):
    """
    讓使用者選擇目前執行計畫 ('產學合作' 或 '研究計畫')。
    選擇完後更新 setting_data['SOURCE']['field']['目前執行計畫']。
    """
    def __init__(self, setting_data, parent=None):
        super().__init__(parent)
        self.setting_data = setting_data
        self.setWindowTitle("計畫目標選擇")
        self.resize(300, 150)

        self.valid_project_aim_options = ["產學合作", "研究計畫"]

        self.label = QLabel("請選擇目前執行計畫:")
        self.combobox = QComboBox()
        self.combobox.addItems(self.valid_project_aim_options)
        # 預設值
        self.combobox.setCurrentIndex(0)

        self.confirm_button = QPushButton("確認")
        self.confirm_button.clicked.connect(self.on_select)

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.combobox)
        layout.addWidget(self.confirm_button)
        self.setLayout(layout)

    def on_select(self):
        current_project_aim = self.combobox.currentText()
        self.setting_data['SOURCE']['field']['目前執行計畫'] = current_project_aim
        QMessageBox.information(self, "成功", f"目前執行計劃目標已經設定為: {current_project_aim}")
        self.accept()  # 結束對話窗並回傳成功訊號


class SelectFileUpdateProjectNameDialog(QDialog):
    """
    讓使用者在 GUI 中選擇對應檔案，並更新設定資料。
    選擇完後同時更新 output 的最終統合檔名。
    """
    def __init__(self, setting_data, parent=None):
        super().__init__(parent)
        self.setting_data = setting_data
        self.setWindowTitle("選擇資料名稱")
        self.resize(400, 200)

        self.current_aim = self.setting_data['SOURCE']['field']['目前執行計畫']

        # 判斷初始資料夾
        if self.current_aim == "研究計畫":
            self.base_directory = os.path.abspath("./data/research_proj/")
        else:
            self.base_directory = os.path.abspath("./data/industry_coop/")

        self.label = QLabel(f"請選擇 {self.current_aim} 的初始資料檔案")
        self.select_button = QPushButton("選擇檔案")
        self.select_button.clicked.connect(self.open_file_dialog)

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.select_button)
        self.setLayout(layout)

        self.selected_file_path = None  # 用來存放使用者選到的檔案路徑

    def open_file_dialog(self):
        file_dialog = QFileDialog(self, f"選擇 {self.current_aim} 的初始資料檔案", directory=self.base_directory)
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                file_path = os.path.abspath(selected_files[0])
                if file_path.startswith(self.base_directory):
                    file_name = os.path.basename(file_path)
                    # 更新 setting_data 中對應的檔名
                    if self.current_aim == '研究計畫':
                        self.setting_data['SOURCE']['data']['research_proj']['研究計畫申請名冊'] = file_name
                    elif self.current_aim == '產學合作':
                        self.setting_data['SOURCE']['data']['industry_coop']['產學合作申請名冊'] = file_name

                    # 更新輸出檔名
                    file_name_only = os.path.splitext(file_name)[0]
                    self.setting_data['OUTPUT']['data']['output']['FINAL_COMMITTEE'] = file_name_only + "_推薦表統合_VBA.xlsx"

                    QMessageBox.information(self, "成功", f"已更新檔案名稱為: {file_name}")
                    self.selected_file_path = file_path
                    self.accept()  # 結束對話窗並回傳成功訊號
                else:
                    QMessageBox.critical(self, "錯誤", "選擇的檔案不在允許的目錄內。")
            else:
                QMessageBox.critical(self, "錯誤", "未選擇任何檔案。")


class SelectSheetFromExcelDialog(QDialog):
    """
    讀取 Excel 檔案中的 Sheet，讓使用者進行多選，更新 setting_data['SOURCE']['field']['計畫SHEET']。
    """
    def __init__(self, file_path, setting_data, parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self.setting_data = setting_data

        self.setWindowTitle("選擇計畫 SHEET (所有 SHEET 欄位記得統一)")
        self.resize(300, 300)

        workbook = load_workbook(self.file_path, read_only=True)
        self.sheets = workbook.sheetnames

        self.label = QLabel("請選擇要使用的 SHEET（可多選）：")

        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for sheet_name in self.sheets:
            item = QListWidgetItem(sheet_name)
            self.list_widget.addItem(item)

        self.confirm_button = QPushButton("確認")
        self.confirm_button.clicked.connect(self.on_select)

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.list_widget)
        layout.addWidget(self.confirm_button)
        self.setLayout(layout)

    def on_select(self):
        selected_items = self.list_widget.selectedItems()
        if selected_items:
            selected_sheets = [item.text() for item in selected_items]
            self.setting_data['SOURCE']['field']['計畫SHEET'] = selected_sheets
            QMessageBox.information(self, "成功", f"計畫 SHEET 已更新為: {', '.join(selected_sheets)}")
            self.accept()  # 結束對話窗
        else:
            QMessageBox.critical(self, "錯誤", "至少選擇一個 SHEET。")


class ConfirmAndUpdateProjectNameColumnDialog(QDialog):
    """
    確認 Excel Sheet 中的欄位，讓使用者選擇計畫相關欄位。
    特別強調「申請主持人欄位」為必填。
    """
    def __init__(self, file_path, sheet_name, setting_data, parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.setting_data = setting_data

        self.setWindowTitle("確認計畫相關欄位")
        self.resize(600, 600)

        workbook = load_workbook(self.file_path, read_only=True)
        sheet = workbook[self.sheet_name]
        self.columns = [cell for cell in next(sheet.iter_rows(max_row=1, values_only=True))]

        # 先讀取 setting 裏的舊值（若有）
        self.current_project_name_col = self.setting_data['SOURCE']['field'].get('計畫名稱', '')
        self.current_keyword_col = self.setting_data['SOURCE']['field'].get('中文關鍵字', '')
        self.current_abstract_col = self.setting_data['SOURCE']['field'].get('計劃摘要', '')
        self.current_personal_title_col = self.setting_data['SOURCE']['field'].get('職稱', '')

        # 做一個可捲動區域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        scroll.setWidget(container)

        self.layout_main = QVBoxLayout(container)
        self.layout_all = QVBoxLayout()
        self._create_form_widgets()

        # 最底下放一個「確認」按鈕
        self.confirm_button = QPushButton("確認")
        self.confirm_button.clicked.connect(self.on_confirm)
        self.layout_main.addWidget(self.confirm_button)
        self.layout_main.addStretch()

        self.layout_all.addWidget(scroll)
        self.setLayout(self.layout_all)

    def _create_form_widgets(self):
        """
        建立整個表單（欄位選擇的下拉、複選清單等）
        """
        # 依序建立各段落
        # 計畫名稱
        label_proj_name = QLabel("請選擇屬於計畫名稱的欄位:")
        self.combo_proj_name = QComboBox()
        self.combo_proj_name.addItem("")
        self.combo_proj_name.addItems(self.columns)
        if self.current_project_name_col in self.columns:
            self.combo_proj_name.setCurrentText(self.current_project_name_col)

        # 中文關鍵字
        label_keyword = QLabel("請選擇屬於中文關鍵字的欄位:")
        self.combo_keyword = QComboBox()
        self.combo_keyword.addItem("")
        self.combo_keyword.addItems(self.columns)
        if self.current_keyword_col in self.columns:
            self.combo_keyword.setCurrentText(self.current_keyword_col)

        # 計劃摘要
        label_abstract = QLabel("請選擇屬於計劃摘要的欄位:")
        self.combo_abstract = QComboBox()
        self.combo_abstract.addItem("")
        self.combo_abstract.addItems(self.columns)
        if self.current_abstract_col in self.columns:
            self.combo_abstract.setCurrentText(self.current_abstract_col)

        # 申請機構
        label_institute = QLabel("請選擇屬於申請機構(學校)的欄位:")
        self.combo_institute = QComboBox()
        self.combo_institute.addItem("")
        self.combo_institute.addItems(self.columns)

        # 主持人(必填)
        label_leader = QLabel("請選擇屬於(計畫)主持人的欄位(必填):")
        self.combo_leader = QComboBox()
        self.combo_leader.addItem("")
        self.combo_leader.addItems(self.columns)

        # 職稱
        label_title = QLabel("請選擇屬於職稱的欄位:")
        self.combo_title = QComboBox()
        self.combo_title.addItem("")
        self.combo_title.addItems(self.columns)
        if self.current_personal_title_col in self.columns:
            self.combo_title.setCurrentText(self.current_personal_title_col)

        # 其他欄位(多選)
        label_others = QLabel("請選擇計畫相關其他欄位 (可複選):")
        self.list_others = QListWidget()
        self.list_others.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for col in self.columns:
            self.list_others.addItem(col)

        # 共同主持人(多選)
        label_co_leader = QLabel("請選擇共同(計畫)主持人的欄位 (可複選):")
        self.list_co_leader = QListWidget()
        self.list_co_leader.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for col in self.columns:
            self.list_co_leader.addItem(col)

        # 共同機構(多選)
        label_co_inst = QLabel("請選擇共同機構(學校)的欄位 (可複選):")
        self.list_co_inst = QListWidget()
        self.list_co_inst.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for col in self.columns:
            self.list_co_inst.addItem(col)

        # 依序放進 layout
        self.layout_main.addWidget(label_proj_name)
        self.layout_main.addWidget(self.combo_proj_name)

        self.layout_main.addWidget(label_keyword)
        self.layout_main.addWidget(self.combo_keyword)

        self.layout_main.addWidget(label_abstract)
        self.layout_main.addWidget(self.combo_abstract)

        self.layout_main.addWidget(label_institute)
        self.layout_main.addWidget(self.combo_institute)

        self.layout_main.addWidget(label_leader)
        self.layout_main.addWidget(self.combo_leader)

        self.layout_main.addWidget(label_title)
        self.layout_main.addWidget(self.combo_title)

        self.layout_main.addWidget(label_others)
        self.layout_main.addWidget(self.list_others)

        self.layout_main.addWidget(label_co_leader)
        self.layout_main.addWidget(self.list_co_leader)

        self.layout_main.addWidget(label_co_inst)
        self.layout_main.addWidget(self.list_co_inst)

    def on_confirm(self):
        proj_name_col = self.combo_proj_name.currentText()
        keyword_col = self.combo_keyword.currentText()
        abstract_col = self.combo_abstract.currentText()
        institute_col = self.combo_institute.currentText()
        leader_col = self.combo_leader.currentText()
        title_col = self.combo_title.currentText()

        selected_others = [item.text() for item in self.list_others.selectedItems()]
        selected_co_leader = [item.text() for item in self.list_co_leader.selectedItems()]
        selected_co_institute = [item.text() for item in self.list_co_inst.selectedItems()]

        # 必填檢查：主持人欄位
        if not leader_col:
            QMessageBox.critical(self, "錯誤", "【申請主持人欄位】不得為空，請選擇一個欄位。")
            return

        # 正常更新設定檔
        self.setting_data['SOURCE']['field']['計畫名稱'] = proj_name_col
        self.setting_data['SOURCE']['field']['中文關鍵字'] = keyword_col
        self.setting_data['SOURCE']['field']['計劃摘要'] = abstract_col
        self.setting_data['SOURCE']['field']['申請機構欄位名稱'] = institute_col
        self.setting_data['SOURCE']['field']['申請主持人欄位名稱'] = leader_col
        self.setting_data['SOURCE']['field']['職稱'] = title_col
        self.setting_data['SOURCE']['field']['計畫相關其他欄位'] = selected_others
        self.setting_data['SOURCE']['field']['申請共同主持人'] = selected_co_leader
        self.setting_data['SOURCE']['field']['申請共同機構欄位名稱'] = selected_co_institute

        msg = (
            f"計畫名稱: {proj_name_col}\n"
            f"中文關鍵字: {keyword_col}\n"
            f"計劃摘要: {abstract_col}\n"
            f"申請機構: {institute_col}\n"
            f"主持人(必填): {leader_col}\n"
            f"職稱: {title_col}\n"
            f"其他欄位: {', '.join(selected_others) if selected_others else '無'}\n"
            f"共同主持人: {', '.join(selected_co_leader) if selected_co_leader else '無'}\n"
            f"共同機構: {', '.join(selected_co_institute) if selected_co_institute else '無'}"
        )
        QMessageBox.information(self, "成功", msg)
        self.accept()


def main():
    app = QApplication(sys.argv)

    try:
        with open(DEFAULT_SETTING_YAML, 'r', encoding='utf-8') as file:
            setting_data = yaml.load(file)

        # 第一步：選擇目前執行計畫
        dlg_aim = SelectProjectAimDialog(setting_data)
        if dlg_aim.exec() != QDialog.DialogCode.Accepted:
            # 若使用者取消或關閉視窗
            return

        # 寫回 yaml
        with open(DEFAULT_SETTING_YAML, 'w', encoding='utf-8') as file:
            yaml.dump(setting_data, file)

        # 第二步：選擇檔案並更新檔案名稱
        dlg_file = SelectFileUpdateProjectNameDialog(setting_data)
        if dlg_file.exec() != QDialog.DialogCode.Accepted:
            return

        with open(DEFAULT_SETTING_YAML, 'w', encoding='utf-8') as file:
            yaml.dump(setting_data, file)

        # 根據目前執行計畫來決定檔案路徑
        current_aim = setting_data['SOURCE']['field']['目前執行計畫']
        if current_aim == "研究計畫":
            file_path = os.path.join(os.path.abspath("./data/research_proj/"),
                                     setting_data['SOURCE']['data']['research_proj']['研究計畫申請名冊'])
        else:
            file_path = os.path.join(os.path.abspath("./data/industry_coop/"),
                                     setting_data['SOURCE']['data']['industry_coop']['產學合作申請名冊'])

        # 第三步：選擇要使用的 Sheet (可多選)
        dlg_sheet = SelectSheetFromExcelDialog(file_path, setting_data)
        if dlg_sheet.exec() != QDialog.DialogCode.Accepted:
            return

        with open(DEFAULT_SETTING_YAML, 'w', encoding='utf-8') as file:
            yaml.dump(setting_data, file)

        # 取得使用者所選擇的 sheet
        selected_sheets = setting_data['SOURCE']['field']['計畫SHEET']
        if not selected_sheets:
            QMessageBox.critical(None, "錯誤", "未選擇任何 SHEET，程式終止。")
            return

        # 第四步：確認並更新計畫相關欄位 (僅以第一個 sheet 為例)
        dlg_columns = ConfirmAndUpdateProjectNameColumnDialog(file_path, selected_sheets[0], setting_data)
        if dlg_columns.exec() != QDialog.DialogCode.Accepted:
            return

        with open(DEFAULT_SETTING_YAML, 'w', encoding='utf-8') as file:
            yaml.dump(setting_data, file)

        QMessageBox.information(None, "完成", "設定已成功更新並存回檔案。")
        app.quit()

    except Exception as e:
        QMessageBox.critical(None, "錯誤", f"發生錯誤，無法更新設定檔。\n{e}")
    finally:
        print("程式結束，請檢查設定檔是否已更新。")
        sys.exit(app.exec())
    

if __name__ == "__main__":
    main()
